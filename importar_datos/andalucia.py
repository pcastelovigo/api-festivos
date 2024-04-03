import os
from openpyxl import Workbook
import dill as pickle
import xlrd
import datetime as dt
from workalendar.europe import Andalusia

from models.holiday import Festivo


def cargar_datos(fichero_excel, year):
	result = set(local_holidays(fichero_excel, year))

	holidays = Andalusia(year=year).holidays()
	for day in holidays:
		init_params = {
			'estado': "es",
			'fecha': day[0].strftime('%Y-%m-%d'),
			'nombre': day[1],
			'autonomia': 'an',
		}
		festivo = Festivo(**init_params)
		result.add(festivo)
	return sorted(list(result))


def local_holidays(fichero_excel, year):
	result = []
	if fichero_excel.name.endswith('.xls'):
		wb = open_excel_without_workbook_part(fichero_excel.path)
		sheet = wb.active

		for row in sheet.iter_rows(min_row=2, values_only=True):
			fecha_dt = dt.datetime.strptime(str(int(row[0])), '%Y%m%d')
			try:
				fecha = fecha_dt.strftime('%Y-%m-%d')
			except AttributeError:
				continue

			init_params = {
				'estado': "es",
				'fecha': fecha,
				'nombre': row[1],
				'provincia': row[3],
				'autonomia': 'an',
				'localidad': row[2],
			}
			festivo = Festivo(**init_params)
			result.append(festivo)
	return result


def open_excel_without_workbook_part(fichero):
	xlrd_workbook = xlrd.open_workbook(fichero)
	xlrd_worksheet = xlrd_workbook.sheet_by_index(0)
	nrows = xlrd_worksheet.nrows
	ncols = xlrd_worksheet.ncols

	openpyxl_book = Workbook()
	openpyxl_worksheet = openpyxl_book.active

	for row in range(0, nrows):
		for col in range(0, ncols):
			openpyxl_worksheet.cell(row=row + 1, column=col + 1).value = xlrd_worksheet.cell_value(row, col)

	return openpyxl_book


if __name__ == "__main__":
	curpath = os.path.abspath(os.curdir)

	for fichero in os.scandir('fuentes/andalucia/'):
		in_year = fichero.name.split('.')[0][-4:]
		festivos = cargar_datos(fichero, in_year)
		print(f'Current path is {curpath}')
		print(f'../datos/{in_year}-es-an.dat')
		with open(f'../datos/{in_year}-es-an.dat', 'w+b') as output_file:
			pickle.dump(festivos, output_file)
